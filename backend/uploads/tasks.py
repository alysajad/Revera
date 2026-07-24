import csv
import io
import re
from datetime import date, datetime

from celery import shared_task
from django.db import transaction

from leads.models import Lead
from .models import UploadBatch, UploadRow
from .storage import download_bytes

SOURCE_MAP = {"meta": Lead.Source.META, "facebook": Lead.Source.META, "instagram": Lead.Source.META, "fb": Lead.Source.META, "website": Lead.Source.WEBSITE, "web": Lead.Source.WEBSITE, "landing page": Lead.Source.WEBSITE, "carwale": Lead.Source.CARWALE, "car wale": Lead.Source.CARWALE, "cw": Lead.Source.CARWALE, "walk-in": Lead.Source.WALKIN, "walk in": Lead.Source.WALKIN, "walkin": Lead.Source.WALKIN}


def value(row, *names):
    lowered = {str(key).strip().lower(): value for key, value in row.items() if key}
    return next((str(lowered[name]).strip() for name in names if lowered.get(name) not in (None, "")), "")


def normalize_phone(phone):
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    if digits.startswith("0") and len(digits) == 11:
        digits = digits[1:]
    return digits if len(digits) == 10 else ""


def classify_source(raw):
    normalized = raw.strip().lower()
    if normalized in SOURCE_MAP:
        return SOURCE_MAP[normalized], raw
    if normalized in {"oem", "telein", "google"}:
        return Lead.Source.OTHER, raw
    if not normalized:
        return Lead.Source.UNKNOWN, raw
    return Lead.Source.CAMPAIGN, raw


def parse_date(raw):
    for format_string in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, format_string).date()
        except ValueError:
            continue
    return date.today()


def read_rows(filename, content):
    if filename.lower().endswith(".csv"):
        yield from csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        return
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, values))


@shared_task
def parse_upload_batch(batch_id):
    batch = UploadBatch.objects.get(id=batch_id)
    try:
        content = download_bytes(batch.storage_path)
        parsed_rows = []
        skipped = 0
        for row_number, row in enumerate(read_rows(batch.filename, content), start=2):
            name = value(row, "name", "customer name")
            phone = normalize_phone(value(row, "phone", "mobile"))
            if not name and not phone:
                continue
            error = "" if name and phone else "Name and phone are required."
            if error:
                skipped += 1
            source, source_label = classify_source(value(row, "source"))
            parsed_rows.append({"row_number": row_number, "phone": phone, "validation_error": error, "data": {"name": name, "email": value(row, "email"), "source": source, "source_label": source_label, "campaign": value(row, "campaign"), "model_interest": value(row, "model", "vehicle interest", "model / vehicle interest"), "city": value(row, "city", "location"), "enquiry_date": parse_date(value(row, "date", "enquiry date")).isoformat()}})
        existing_leads = {}
        for lead in Lead.objects.filter(phone__in={row["phone"] for row in parsed_rows if row["phone"]}, deleted_at__isnull=True).only("id", "phone").order_by("id"):
            existing_leads.setdefault(lead.phone, lead)
        staged = [UploadRow(batch=batch, row_number=row["row_number"], normalized_phone=row["phone"], validation_error=row["validation_error"], duplicate_of=(duplicate := existing_leads.get(row["phone"])), resolution=UploadRow.Resolution.PENDING if duplicate else UploadRow.Resolution.IMPORT, data=row["data"]) for row in parsed_rows]
        duplicates = sum(row.duplicate_of_id is not None for row in staged)
        with transaction.atomic():
            UploadRow.objects.filter(batch=batch).delete()
            UploadRow.objects.bulk_create(staged)
            batch.total_rows = len(staged)
            batch.parsed_ok = len([row for row in staged if not row.validation_error])
            batch.duplicates_found = duplicates
            batch.skipped = skipped
            batch.status = UploadBatch.Status.READY
            batch.save(update_fields=["total_rows", "parsed_ok", "duplicates_found", "skipped", "status"])
    except Exception as error:
        batch.status = UploadBatch.Status.FAILED
        batch.error_message = str(error)[:1000]
        batch.save(update_fields=["status", "error_message"])
        raise
